from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from selenium import webdriver
from time import sleep
f = open('URL.txt','r')
wb = Workbook()
ws = wb.active
ws['A1'] = 'Phone'
URL = f.read()
HEADERS = {'User-Agent' : 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 YaBrowser/20.7.1.68 Yowser/2.5 Safari/537.36'}
response = requests.get(URL,headers = HEADERS)
soup = BeautifulSoup(response.content,'lxml')
pages = int(soup.find('div',class_='pager rel clr').text.split()[-3])
i = 1
x = 2
links = []
y = 5
for i in range(pages):
	URL = URL + '?page=' + str(i)
	HEADERS = {'User-Agent' : 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 YaBrowser/20.7.1.68 Yowser/2.5 Safari/537.36'}
	response = requests.get(URL,headers = HEADERS)
	soup = BeautifulSoup(response.content,'lxml')
	items = soup.findAll('tr',class_='wrap')
	for item in items:
		links.append(item.find('a').get('href'))
driver = webdriver.PhantomJS()### PhantomJS  Chrome
for link in links:
	try:
		driver.get(link)
		button = driver.find_element_by_xpath('//span[@class="button inverted spoiler"]').click()
		sleep(1)
		phone = driver.find_element_by_xpath('//div[@class="contactitem"]').text
		while phone[-1] == 'x':
			driver.get(link)
			button = driver.find_element_by_xpath('//span[@class="button inverted spoiler"]').click()
			sleep(3)
			phone = driver.find_element_by_xpath('//div[@class="contactitem"]').text
			if phone[-1] == 'x':
				driver.get(link)
				button = driver.find_element_by_xpath('//span[@class="button inverted spoiler"]').click()
				sleep(30)
				phone = driver.find_element_by_xpath('//div[@class="contactitem"]').text
				if phone[-1] == 'x':
					print('1')
					continue
		if phone[0] == '3':
			phone = '+' + phone
		elif phone[0] == '0':
			phone = '+38' + phone
		else:
			phone = phone
		ws['A' + str(x)] = phone
		print(x)
		x+=1
	except:
		continue
print('Готово!')
wb.save("olx.xlsx")