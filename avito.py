from bs4 import BeautifulSoup
import requests
import pytesseract
from openpyxl import Workbook
from selenium import webdriver
from time import sleep
from PIL import Image
pytesseract.pytesseract.tesseract_cmd = 'C:\\Program Files\\Tesseract-OCR\\tesseract.exe'
wb = Workbook()
ws = wb.active
names = []
links = []
prices = []
Phone = []
ws['A1'] = 'Name'
ws['B1'] = 'Link'
ws['C1'] = 'Price'
ws['D1'] = 'Phone'
a = 1
b = 1
c = 1
d = 1
###
URL = 'https://www.avito.ru/leningradskaya_oblast/avtomobili?p='
HEADERS = {
'User-Agent' : 
'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 YaBrowser/20.7.1.68 Yowser/2.5 Safari/537.36'
}
response =requests.get(URL, headers = HEADERS)
soup = BeautifulSoup(response.content, 'html.parser')
#узнаём колл-во страниц на сайте
pages = int(soup.find('div' , class_ = 'pagination-root-2oCjZ').text.split('...')[-1].split('След')[0])
driver = webdriver.Chrome()
for i in range(pages):
	items = soup.findAll('div' , class_ = 'snippet-horizontal')
	for item in items:
		links.append('https://www.avito.ru' + item.find('a' , class_ = 'snippet-link').get('href'))
		names.append(item.find('a' , class_ = 'snippet-link').text)
		prices.append(item.find('span',class_='snippet-price').text.split('₽')[0])
for name in names:
	a = a + 1
	ws['A' + str(a)] = name
for link in links:
	b = b + 1
	ws['B' + str(b)] = link
for price in prices:
	c = c + 1
	ws['C' + str(c)] = price
for link in links:
	try:
		driver.get(link)
		button = driver.find_element_by_xpath('//a[@class="button item-phone-button js-item-phone-button button-origin contactBar_greenColor button-origin_full-width button-origin_large-extra item-phone-button_hide-phone item-phone-button_card js-item-phone-button_card contactBar_height"]')
		button.click()
		sleep(3);
		driver.save_screenshot('screenshot.png')
		image = Image.open('screenshot.png')
		img_phone = driver.find_element_by_xpath('//div[@class="item-phone-big-number js-item-phone-big-number"]//*')
		location = img_phone.location
		size =img_phone.size
		x = location['x']
		y = location['y']
		width = size['width']
		height = size['height']
		cropped = image.crop((x,y,x + width,y + height))
		cropped.save('phone.png')
		img = Image.open('phone.png')
		phone = pytesseract.image_to_string(img)
		if len(phone) == 15:
			ws['D' + str(d)] = phone
			d = d + 1
		else:
			ws['D' + str(d)] = 'нет'
			d = d + 1
	except:
		try:
			driver.get(link)
			button = driver.find_element_by_xpath('//a[@class="button item-phone-button js-item-phone-button button-origin contactBar_greenColor button-origin_full-width button-origin_large-extra item-phone-button_hide-phone item-phone-button_card js-item-phone-button_card contactBar_height"]')
			button.click()
			sleep(3);
			driver.save_screenshot('screenshot.png')
			image = Image.open('screenshot.png')
			img_phone = driver.find_element_by_xpath('//div[@class="item-phone-big-number js-item-phone-big-number"]//*')
			location = img_phone.location
			size =img_phone.size
			x = location['x']
			y = location['y']
			width = size['width']
			height = size['height']
			cropped = image.crop((x,y,x + width,y + height))
			cropped.save('phone.png')
			img = Image.open('phone.png')
			phone = pytesseract.image_to_string(img)
			if len(phone) == 15:
				ws['D' + str(d)] = phone
				d = d + 1
			else:
				ws['D' + str(d)] = 'нет'
				d = d + 1
		except:
			ws['D' + str(d)] = 'нет'
			d = d + 1

wb.save("avito.xlsx")






